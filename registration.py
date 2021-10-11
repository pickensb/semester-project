from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('C:\\Users\\Admin\\Desktop\\Dance Registration.xlsx')

# create the sheet object
sheet = wb.active


def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 30
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 50
    sheet.column_dimensions['G'].width = 50

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Age"
    sheet.cell(row=1, column=3).value = "Class Type: Combo/Single"
    sheet.cell(row=1, column=4).value = "Day Available"
    sheet.cell(row=1, column=5).value = "Payment: Cash/Card"
    sheet.cell(row=1, column=6).value = "Email"
    sheet.cell(row=1, column=7).value = "Phone Number"


# Function to set focus (cursor)
def focus1(event):
    # set focus on the course_field box
    age_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the sem_field box
    classType_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the form_no_field box
    dayAvailable_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the contact_no_field box
    payment_no_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the email_id_field box
    email_id_field.focus_set()


# Function to set focus
def focus6(event):
    # set focus on the address_field box
    phoneNum_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    name_field.delete(0, END)
    age_field.delete(0, END)
    classType_field.delete(0, END)
    dayAvailable_field.delete(0, END)
    payment_no_field.delete(0, END)
    email_id_field.delete(0, END)
    phoneNum_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
    # if user not fill any entry
    # then print "empty input"
    if (name_field.get() == "" and
            age_field.get() == "" and
            classType_field.get() == "" and
            dayAvailable_field.get() == "" and
            payment_no_field.get() == "" and
            email_id_field.get() == "" and
            phoneNum_field.get() == ""):

        print("empty input")

    else:

        # assigning the max row and max column
        # value up to which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = age_field.get()
        sheet.cell(row=current_row + 1, column=3).value = classType_field.get()
        sheet.cell(row=current_row + 1, column=4).value = dayAvailable_field.get()
        sheet.cell(row=current_row + 1, column=5).value = payment_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = phoneNum_field.get()

        # save the file
        wb.save('C:\\Users\\Admin\\Desktop\\excel.xlsx')

        # set focus on the name_field box
        name_field.focus_set()

        # call the clear() function
        clear()


# Driver code
if __name__ == "__main__":
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light pink')

    # set the title of GUI window
    root.title("registration form")

    # set the configuration of GUI window
    root.geometry("500x300")

    excel()

    # create a Form label
    heading = Label(root, text="Dance Registration", bg="light pink")

    # create a Name label
    name = Label(root, text="Name: ", bg="light pink")

    # create a Course label
    age = Label(root, text="Age: ", bg="light pink")

    # create a Semester label
    classType = Label(root, text="Class Type: Combo/Single", bg="light pink")

    # create a Form No. label
    dayAvailable = Label(root, text="Day of Class:", bg="light pink")

    # create a Contact No. label
    payment_no = Label(root, text="Payment: Cash/Card", bg="light pink")

    # create a Email id label
    email_id = Label(root, text="Email: ", bg="light pink")

    # create a address label
    phoneNum = Label(root, text="Phone Number: ", bg="light pink")

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    age.grid(row=2, column=0)
    classType.grid(row=3, column=0)
    dayAvailable.grid(row=4, column=0)
    payment_no.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    phoneNum.grid(row=7, column=0)

    # create a text entry box
    # for typing the information
    name_field = Entry(root)
    age_field = Entry(root)
    classType_field = Entry(root)
    dayAvailable_field = Entry(root)
    payment_no_field = Entry(root)
    email_id_field = Entry(root)
    phoneNum_field = Entry(root)

    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    name_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    age_field.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    classType_field.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    daysAvailable_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus5 function
    payment_no_field.bind("<Return>", focus5)

    # whenever the enter key is pressed
    # then call the focus6 function
    email_id_field.bind("<Return>", focus6)

    # whenever the enter key is pressed
    # then call the focus6 function
    phoneNum_id_field.bind("<Return>", focus6)

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    name_field.grid(row=1, column=1, ipadx="100")
    age_field.grid(row=2, column=1, ipadx="100")
    classType_field.grid(row=3, column=1, ipadx="100")
    daysAvailable_field.grid(row=4, column=1, ipadx="100")
    payment_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")
    phoneNum_field.grid(row=7, column=1, ipadx="100")

    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                    bg="Red", command=insert)
    submit.grid(row=8, column=1)

    # start the GUI
    root.mainloop()